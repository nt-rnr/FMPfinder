from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import Workbook, load_workbook
import time 
from getpass import getpass



service = Service(executable_path = "chromedriver.exe")
driver = webdriver.Chrome(service = service)
book = openpyxl.load_workbook('C:/Users/morir/OneDrive/Documents/FacebookMPscript/prices.xlsx')
sheet = book.active

# user = input('what is you username: ')
# passw = getpass('What is your password: ')
product = input('what product do you want: ')
product = product.replace(" ", "+")
minPrice = input('What is the minimum price: ')
maxPrice = input('What is the maximum price: ')
city = input('in what city: ')
totalPrice = 0
numOfProducts = 0
lowest = 999999999
lowestLink = ""
g = 1
cell = ''

driver.get("http://facebook.com/marketplace/"+city+"/search/?query="+product)
driver.maximize_window() 

WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Close']"))
    )

close_button = driver.find_element(By.XPATH, "//div[@aria-label='Close']")
close_button.click()



#if you want to enter your username (might bug scrolling)
# Username = driver.find_element(By.XPATH, "//input[@dir='ltr']")
# Username.send_keys(user)
# Password = driver.find_element(By.XPATH, "//input[@name='pass']")
# Password.send_keys(passw)
# loginButton = driver.find_element(By.XPATH, "//div[@aria-label='Log In']")
# loginButton.click()
# WebDriverWait(driver, 10).until(
#     EC.presence_of_element_located((By.CSS_SELECTOR, 'span.x1lliihq.x6ikm8r.x10wlt62.x1n2onr6.xlyipyv.xuxw1ft.x1j85h84'))
#     )

SCROLL_PAUSE_TIME = 5




for i in range(3):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(SCROLL_PAUSE_TIME)  
      
elements = driver.find_elements(By.CSS_SELECTOR, 'span.x1lliihq.x6ikm8r.x10wlt62.x1n2onr6.xlyipyv.xuxw1ft.x1j85h84')
priceLoc = driver.find_elements(By.CSS_SELECTOR, 'span.x193iq5w.xeuugli.x13faqbe.x1vvkbs.x1xmvt09.x1lliihq.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.xudqn12.x676frb.x1lkfr7t.x1lbecb7.x1s688f.xzsf02u')
links = driver.find_elements(By.CSS_SELECTOR, 'a.x1i10hfl.xjbqb8w.x1ejq31n.xd10rxx.x1sy0etr.x17r0tee.x972fbf.xcfux6l.x1qhh985.xm0m39n.x9f619.x1ypdohk.xt0psk2.xe8uvvx.xdj266r.x11i5rnm.xat24cr.x1mh8g0r.xexx8yu.x4uap5.x18d9i69.xkhd6sd.x16tdsg8.x1hl2dhg.xggy1nq.x1a2a7pz.x1heor9g.x1sur9pj.xkrqix3.x1lku1pv')

for i, (value1, value2, value3) in enumerate(zip(elements, priceLoc, links)):
    location = value1.text
    price = value2.text   
    cleanedPrice = price.strip("CA$").replace(",", "").strip()
    if(cleanedPrice.isnumeric() and int(cleanedPrice) != 0 and int(cleanedPrice) != 1 and int(cleanedPrice) >= int(minPrice) and int(cleanedPrice) <= int(maxPrice)): 
        if(lowest > int(cleanedPrice)):
            lowest = int(cleanedPrice)
            lowestLink = value3.get_attribute('href')
        totalPrice += int(cleanedPrice)
        numOfProducts += 1


while(True):
    if(sheet.cell(row=g, column=1).value == None):
        sheet.cell(row=g, column=1).value =  numOfProducts
        sheet.cell(row=g, column=2).value = totalPrice/numOfProducts
        sheet.cell(row=g, column=3).value = lowest
        sheet.cell(row=g, column=4).value = lowestLink
        sheet.cell(row=g, column=4).hyperlink = lowestLink
        break
    g += 1

book.save('C:/Users/morir/OneDrive/Documents/FacebookMPscript/prices.xlsx')

print(numOfProducts)
print(totalPrice/numOfProducts)
print("Lowest price is: ", lowest)  
print(lowestLink)
time.sleep(15)
